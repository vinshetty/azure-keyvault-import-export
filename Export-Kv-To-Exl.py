import os
import cmd
import datetime
import argparse
from azure.keyvault.secrets import SecretClient
from azure.identity import DefaultAzureCredential,AzureCliCredential
from tabulate import tabulate
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl import Workbook

parser = argparse.ArgumentParser()
parser.add_argument("-kv", "--KeyVault", help = "Provide KV name which want to see the value")
# Read arguments from command line
args = parser.parse_args()
# Read arguments from command line
args = parser.parse_args()

keyVaultName = args.KeyVault
#keyVaultName = os.environ["KEY_VAULT_NAME"]

KVUri = f"https://{keyVaultName}.vault.azure.net"
#
wb = Workbook()
ws = wb.create_sheet(index=0, title=keyVaultName)
ws.cell(row=1, column=1).value = 'Key'
ws.cell(row=1, column=2).value = 'value'
r = 2
printresults = []
headers = ["SecretName", "Value"]

try:
   credential = AzureCliCredential()
   secret_client = SecretClient(vault_url=KVUri, credential=credential)
   secret_properties = secret_client.list_properties_of_secrets()
   for secret_property in secret_properties:
          secretcheck = (secret_property.enabled)
          #print(secretcheck)
          if secretcheck == True:
              secret = secret_client.get_secret(secret_property.name)
              ws.cell(row=r, column=1).value = secret.name
              ws.cell(row=r, column=2).value = secret.value
              r += 1
              printresults.append([secret.name, secret.value])
          else:
              print(secret_property.name + " This seceret is disabled")
except Exception as e:
  print(e)  
  print("Something went wrong", type(e).__name__)
  pass
date = datetime.datetime.now().date()
wb.save(str(keyVaultName)+'-'+str(date)+'.xlsx')
print(tabulate(printresults,headers,tablefmt="simple"))
